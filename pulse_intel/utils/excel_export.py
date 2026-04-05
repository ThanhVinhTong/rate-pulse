import os
import re
from datetime import datetime, timezone
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


MAX_COLUMN_WIDTH = 80
GEO_COUNTS_PATTERN = re.compile(r"(\d+)\s+signal types?.*?(\d+)\s+events?", re.IGNORECASE)
SHEET_SAFE_PATTERN = re.compile(r"[\[\]\*\?/\\:]")


def _extract_region_name(raw_label: str) -> str:
    first_line = raw_label.split("\n", 1)[0].strip()
    cleaned = re.sub(r"^[^\w]+", "", first_line).strip()
    return cleaned or first_line


def _build_geo_rows(ai_insights: dict) -> list[dict]:
    rows: list[dict] = []
    geo_insights = ai_insights.get("geo_insights", [])
    if not isinstance(geo_insights, list):
        return rows

    for insight in geo_insights:
        if not isinstance(insight, dict):
            continue
        for label, detail in insight.items():
            matched = GEO_COUNTS_PATTERN.search(label)
            signal_types = int(matched.group(1)) if matched else ""
            events = int(matched.group(2)) if matched else ""
            rows.append(
                {
                    "region": _extract_region_name(str(label)),
                    "summary": str(label).replace("\n", " | "),
                    "signal_types": signal_types,
                    "events": events,
                    "detail": str(detail),
                }
            )
    return rows


def _build_breaking_news_rows(ai_insights: dict) -> list[dict]:
    rows: list[dict] = []
    breaking_news = ai_insights.get("break_news", [])
    if not isinstance(breaking_news, list):
        return rows

    for index, headline in enumerate(breaking_news, start=1):
        rows.append({"index": index, "headline": str(headline)})
    return rows


def _build_feed_rows(category: str, section_data: dict) -> list[dict]:
    rows: list[dict] = []
    if not isinstance(section_data, dict):
        return rows

    for title, payload in section_data.items():
        href = ""
        published = ""
        if isinstance(payload, dict):
            href = str(payload.get("href", ""))
            published = str(payload.get("time", ""))

        rows.append(
            {
                "category": category,
                "title": str(title),
                "time": published,
                "url": href,
                "domain": urlparse(href).netloc if href else "",
            }
        )
    return rows


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, MAX_COLUMN_WIDTH)


def _write_rows(sheet: Worksheet, headers: list[str], rows: list[dict]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)


def _get_sheet_name(raw_name: str, used_names: set[str]) -> str:
    cleaned = SHEET_SAFE_PATTERN.sub("_", raw_name).strip() or "sheet"
    base = cleaned[:31]
    candidate = base
    index = 1
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def export_news_to_excel(news_data: dict, output_folder: str, filename: str | None = None) -> str:
    os.makedirs(output_folder, exist_ok=True)
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pulse_intel_{timestamp}.xlsx"

    ai_insights = news_data.get("ai_insights", {})
    if not isinstance(ai_insights, dict):
        ai_insights = {}

    feed_sections = {
        key: value
        for key, value in news_data.items()
        if key != "ai_insights" and isinstance(value, dict)
    }

    all_feed_rows: list[dict] = []
    for category, section_data in feed_sections.items():
        all_feed_rows.extend(_build_feed_rows(category, section_data))

    overview_rows = [
        {
            "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "world_brief": ai_insights.get("world_brief", ""),
            "geo_insight_count": len(ai_insights.get("geo_insights", []))
            if isinstance(ai_insights.get("geo_insights", []), list)
            else 0,
            "breaking_news_count": len(ai_insights.get("break_news", []))
            if isinstance(ai_insights.get("break_news", []), list)
            else 0,
            "feed_category_count": len(feed_sections),
            "feed_item_count": len(all_feed_rows),
        }
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_names: set[str] = set()

    overview_sheet = workbook.create_sheet(_get_sheet_name("overview", used_sheet_names))
    _write_rows(
        overview_sheet,
        [
            "generated_at_utc",
            "world_brief",
            "geo_insight_count",
            "breaking_news_count",
            "feed_category_count",
            "feed_item_count",
        ],
        overview_rows,
    )

    geo_sheet = workbook.create_sheet(_get_sheet_name("geo_insights", used_sheet_names))
    _write_rows(
        geo_sheet,
        ["region", "summary", "signal_types", "events", "detail"],
        _build_geo_rows(ai_insights),
    )

    breaking_news_sheet = workbook.create_sheet(_get_sheet_name("breaking_news", used_sheet_names))
    _write_rows(
        breaking_news_sheet,
        ["index", "headline"],
        _build_breaking_news_rows(ai_insights),
    )

    feeds_all_sheet = workbook.create_sheet(_get_sheet_name("feeds_all", used_sheet_names))
    _write_rows(
        feeds_all_sheet,
        ["category", "title", "time", "url", "domain"],
        all_feed_rows,
    )

    for category, section_data in feed_sections.items():
        category_rows = _build_feed_rows(category, section_data)
        category_sheet = workbook.create_sheet(_get_sheet_name(category, used_sheet_names))
        _write_rows(
            category_sheet,
            ["title", "time", "url", "domain"],
            [
                {
                    "title": row["title"],
                    "time": row["time"],
                    "url": row["url"],
                    "domain": row["domain"],
                }
                for row in category_rows
            ],
        )

    output_path = os.path.join(output_folder, filename)
    workbook.save(output_path)
    return output_path
